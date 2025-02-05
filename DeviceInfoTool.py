import time
import os
import keyboard
import requests
from requests.auth import HTTPBasicAuth
import socket 
import concurrent.futures
import pandas as bamboo
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from concurrent.futures import ThreadPoolExecutor, as_completed
from playwright.sync_api import sync_playwright
from datetime import datetime
from bs4 import BeautifulSoup
from ping3 import ping

"""
An example program that scans various subnets on a network and grabs info about all fileservers and routers, as well as their firmware and other important info, then saves
everything to an Excel spreadsheet.
"""

#subnetPing and subnetScanning are utilized to grab all ip addresses on the subnet and store them in a tuple.

def subnetPing(subnetIP):
    if ping(subnetIP, timeout = 1):
        try:
            nameDevice = socket.gethostbyaddr(subnetIP)[0]
        except socket.herror:
            nameDevice = None
        
        return subnetIP, nameDevice
    
    return None

def subnetScanning(subnetInput):
    ipAddressesFound = [f"{subnetInput}{i}" for i in range(1, 255)]
    
    with concurrent.futures.ThreadPoolExecutor(max_workers = 50) as executor:
        resultsList = list(executor.map(subnetPing, ipAddressesFound))
    
    return [result for result in resultsList if result is not None]

#Router Firmware Check    
def routerFirmwareChecker(ipInput):
    ipFound = ipInput[0]
    deviceIPFound = ipInput[1]
    locationFound = ipInput[2]
    
    if ipFound != "1.1.1.1":
        appendedIP = "http://" + ipFound
        
        response = requests.get(appendedIP)
        
        if response.status_code == 200:
            
            htmlSoup = BeautifulSoup(response.content, 'html.parser')

            locationDiv = htmlSoup.find("div", id = "critRouterInfo")
            
            if locationDiv:

                locationTable = locationDiv.find("table")

                tableRows = locationTable.find_all("tr")

                secondRow = tableRows[1]
                
                secondInnerRow = secondRow.find_all("td")[1].text

                firstRow = tableRows[0]
                
                firstInnerRow = firstRow.find_all("td")[1].text
            
            else:
                locationDiv = htmlSoup.find("div", id = "routerSettings")
                locationTable = locationDiv.find("table")
                tableRows = locationTable.find_all("tr")
                
                secondRow = tableRows[1]
                
                secondInnerRow = secondRow.find_all("td")[1].text

                firstRow = tableRows[0]
                
                firstInnerRow = firstRow.find_all("td")[1].text
        
        #Some issue prevented data from being recovered properly.
        
        else:
            secondInnerRow, firstInnerRow = "", ""
        
        #Return in this order: Firmware version, MAC address, IP address, deviceIPFound, locationFound 
        return secondInnerRow, firstInnerRow, ipFound, deviceIPFound, locationFound        

#File server firmware checking functions.    

#First function covers some of the file servers that are easily accessible to grab info from without javascript.
def fileServerSecondaryFirmwareChecker(ipInput):
    deviceFound = ipInput[0]
    ipFound = ipInput[1]
    locationFound = ipInput[2]

    if deviceFound != "1.1.1.1":
        appendedIP = "http://" + deviceFound
        
        response = requests.get(appendedIP)

        if response.status_code == 200:                    
            htmlSoup = BeautifulSoup(response.content, 'html.parser')
            iframe = htmlSoup.find("iframe", {"name": "mainFilePage"})
            
            iframeURL = appendedIP + iframe["srcFiles"]
            
            iframeResponse = requests.get(iframeURL)
            
            if response.status_code == 200:
                iframeSoup = BeautifulSoup(iframeResponse.content, 'html.parser')
                
                mainBody = iframeSoup.find("body")
                
                table2 = mainBody.find("table", id = "Table2")
                
                tableRows = table2.find_all("tr")
                
                firmware = tableRows[2]
                
                separatedFirmware = firmware.find_all("td")[1].text if firmware.find_all("td")[1] else "y"

                macAddress = tableRows[4]
                
                separatedMacAddress = macAddress.find_all("td")[1].text if macAddress.find_all("td")[1] else "y"
                
                newFileServerFirmware, newFileServerBootFirmware, newFileServerIndicator = "", "", "N"
                
            else:
                separatedMacAddress, separatedFirmware, newFileServerFirmware, newFileServerBootFirmware, newFileServerIndicator = "", "", "", "", ""
        
        return deviceFound, ipFound, locationFound, separatedFirmware, separatedMacAddress, newFileServerFirmware, newFileServerBootFirmware, newFileServerIndicator

#This second function covers some newer network file servers with greater security measures implemented.
def newFileServerFirmwareHelper(deviceNameURL):
    username = "Neo"
    password = "FollowTheWhiteRabbit1999"
    
    authenticatedURL = f"http://{username}:{password}@{deviceNameURL.strip('http://')}"
    
    response = requests.get(deviceNameURL, auth = HTTPBasicAuth(username, password))

    #If "Authentication - Access Denied." not in response.text:
    
    if response.status_code != 200:
        
        with sync_playwright() as play:
            edgeLocation = "C:\\Program Files\\Microsoft\\Edge\\Application\\msedge.exe"
            agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edge/126.0.0.0"
            webBrowser = play.chromium.launch(executable_path = edgeLocation, headless = True, args=["--no-sandbox", "--disable-blink-features=AutomationControlled", "--headless=new"])
            context = webBrowser.new_context(user_agent = agent)
            page = context.new_page()
            
            try:
                
                page.goto(authenticatedURL, timeout = 30000)
                
                page.wait_for_load_state('networkidle')
                
                page.click('button:has-text("Network Info")')
                
                page.wait_for_load_state('networkidle')
                
                page.wait_for_selector("input#mac")
                
                macAddressFileServer = page.input_value("input#mac")
                
                page.click('a:has-text("Server Management Info")')
                
                page.wait_for_load_state('networkidle')
                
                soup = BeautifulSoup(page.content(), 'html.parser')
                
                fileNetworkFirmware = soup.find("label", id = "firmVersion").text
                
                fileServerSecondaryFirmware = soup.find("label", id = "secondaryFirm").text
                
                newFileServerBootFirmware = soup.find("label", id = "newBootVersion").text
                
                webBrowser.close()
                
                return macAddressFileServer, fileNetworkFirmware, fileServerSecondaryFirmware, newFileServerBootFirmware, "Y"
                
            except:
                webBrowser.close()
                print(f"Something went wrong for {deviceNameURL}")
                return "", "", "", "", "Y"

    else:
        print(f"Failed to fetch the requested webpage: {deviceNameURL}. Generated the following response code: {response.status_code}")
        return "", "", "", "", "Y"

def excelFirmwareAppend(localFilePath, sharedFilePath, routerTopRow, fileServerTopRow, inputTupleListRouter, inputTupleListFileServer):
    #Create the excel docs if they don't already exist
    if not (os.path.exists(localFilePath)):
        with bamboo.ExcelWriter(localFilePath, engine = "openpyxl") as excelWrite:
            firstRowDataFrameRouter = bamboo.DataFrame(columns = routerTopRow)
            
            firstRowDataFrameRouter.to_excel(excelWrite, sheet_name = "Routers", index = False)
            
            firstRowDataFrameFileServer = bamboo.DataFrame(columns = fileServerTopRow)
            
            firstRowDataFrameFileServer.to_excel(excelWrite, sheet_name = "FileServers", index = False)

    if not (os.path.exists(sharedFilePath)):
        with bamboo.ExcelWriter(sharedFilePath, engine = "openpyxl") as excelWrite:
            firstRowDataFrameRouter = bamboo.DataFrame(columns = routerTopRow)
            
            firstRowDataFrameRouter.to_excel(excelWrite, sheet_name = "Routers", index = False)
            
            firstRowDataFrameFileServer = bamboo.DataFrame(columns = fileServerTopRow)
            
            firstRowDataFrameFileServer.to_excel(excelWrite, sheet_name = "FileServers", index = False)

    #Append data to the local excel document
    with bamboo.ExcelWriter(localFilePath, engine = "openpyxl", mode = "a", if_sheet_exists = "overlay") as excelWrite:
        #Append scanner info to local excel document and save
        existingExcelDataFrameLocalRouter = bamboo.read_excel(localFilePath, sheet_name = "Routers", engine = "openpyxl")
        remainderExcelDataFrameLocalRouter = bamboo.DataFrame(inputTupleListRouter, columns = routerTopRow)
        existingExcelDataFrameLocalRouter = bamboo.concat([existingExcelDataFrameLocalRouter, remainderExcelDataFrameLocalRouter], ignore_index = True)
        existingExcelDataFrameLocalRouter.to_excel(excelWrite, sheet_name = "Routers", index = False)
        
        #Append printer info to local excel document and save
        if inputTupleListFileServer[0][0] != " ":
            existingExcelDataFrameLocalFileServer = bamboo.read_excel(localFilePath, sheet_name = "FileServers", engine = "openpyxl")
            remainderExcelDataFrameLocalFileServer = bamboo.DataFrame(inputTupleListFileServer, columns = fileServerTopRow)
            existingExcelDataFrameLocalFileServer = bamboo.concat([existingExcelDataFrameLocalFileServer, remainderExcelDataFrameLocalFileServer], ignore_index = True)
            existingExcelDataFrameLocalFileServer.to_excel(excelWrite, sheet_name = "FileServers", index = False)

    #Append data to the shared excel document
    with bamboo.ExcelWriter(sharedFilePath, engine = "openpyxl", mode = "a", if_sheet_exists = "overlay") as excelWrite:
        #Append scanner info to local excel document and save
        existingExcelDataFrameSharedRouter = bamboo.read_excel(sharedFilePath, sheet_name = "Routers", engine = "openpyxl")
        remainderExcelDataFrameSharedRouter = bamboo.DataFrame(inputTupleListRouter, columns = routerTopRow)
        existingExcelDataFrameSharedRouter = bamboo.concat([existingExcelDataFrameSharedRouter, remainderExcelDataFrameSharedRouter], ignore_index = True)
        existingExcelDataFrameSharedRouter.to_excel(excelWrite, sheet_name = "Routers", index = False)
        
        #Append printer info to local excel document and save
        if inputTupleListFileServer[0][0] != " ":
            existingExcelDataFrameSharedFileServer = bamboo.read_excel(sharedFilePath, sheet_name = "FileServers", engine = "openpyxl")
            remainderExcelDataFrameSharedFileServer = bamboo.DataFrame(inputTupleListFileServer, columns = fileServerTopRow)
            existingExcelDataFrameSharedFileServer = bamboo.concat([existingExcelDataFrameSharedFileServer, remainderExcelDataFrameSharedFileServer], ignore_index = True)
            existingExcelDataFrameSharedFileServer.to_excel(excelWrite, sheet_name = "FileServers", index = False)

    #Autofit the local excel document column widths
    localWorkbook = load_workbook(localFilePath)
    localRouterSheet = localWorkbook["Routers"]
    localFileServerSheet = localWorkbook["FileServers"]
    
    autofitColumnWidth(localRouterSheet)
    autofitColumnWidth(localFileServerSheet)
    
    localWorkbook.save(localFilePath)
    
    #Autofit the shared excel document column widths
    sharedWorkbook = load_workbook(sharedFilePath)
    sharedRouterSheet = sharedWorkbook["Routers"]
    sharedFileServerSheet = sharedWorkbook["FileServers"]
    
    autofitColumnWidth(sharedRouterSheet)
    autofitColumnWidth(sharedFileServerSheet)
    
    sharedWorkbook.save(sharedFilePath)

def autofitColumnWidth(inputSheet):
    for column in inputSheet.columns:
        maxLength = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > maxLength:
                    maxLength = len(str(cell.value))
            except:
                pass
        
        adjustedWidth = (maxLength + 4)
        inputSheet.column_dimensions[column_letter].width = adjustedWidth

timeStart = time.time()

print("Please wait about 12 seconds for all scanner values to download...\n")

#Below read-in from the txt file grabs subnets/subnet locations and stores results into lists

#Get current date and time
estDateTime = datetime.now()

estDate = estDateTime.strftime("%Y-%m-%d")
estTime = estDateTime.strftime("%H:%M:%S")

estDateTimeTuple = (estDate, estTime)

subnetsReadIn = []
locationsReadIn = []
browserInstance = None

with open("Subnet.txt", "r") as fReader:
    lineReadIn = fReader.readlines()
    
    for lineRead in lineReadIn:
        lineRead = lineRead.strip()
        
        if "|" in lineRead:
            subnetRead, locationRead = lineRead.split("|", 1)
            
            subnetRead = subnetRead.strip()
            locationRead = locationRead.strip()
            
            subnetsReadIn.append(subnetRead)
            locationsReadIn.append(locationRead)

#Below for loop is utilized to grab all "router" (Routers) and "fileserver" (FileServers) on the subnets and stores them in tuples.

routerDevicesFoundTuple = []
fileserverDevicesFoundTuple = []

for i in range(len(subnetsReadIn)):
    devicesFound = subnetScanning(subnetsReadIn[i])

    for ip, device_name in devicesFound:
        if (device_name and device_name.startswith("router")) or (device_name and device_name.startswith("ROUTER")):
            routerDevicesFoundTuple.append((device_name, ip, locationsReadIn[i]))
        
        elif (device_name and device_name.startswith("fileserver")) or (device_name and device_name.startswith("FILESERVER")):
            fileserverDevicesFoundTuple.append((device_name, ip, locationsReadIn[i]))

firmwareTuple = []
firmwareMacTuple = []

#Multithreading utilized to speed up javascript parsing operations through playwright, at the cost of temporarily higher CPU usage.
with ThreadPoolExecutor(max_workers = 50) as executor:
    futureTuple = {executor.submit(routerFirmwareChecker, foundTuple): foundTuple for foundTuple in routerDevicesFoundTuple}
    
    for future in as_completed(futureTuple):
        foundTuple = futureTuple[future]
        
        try:
            firmware, macAddress, deviceFound, ipFound, locationFound = future.result()
            
            newTuple = (deviceFound, ipFound, locationFound, firmware, macAddress) + estDateTimeTuple
            firmwareMacTuple.append(newTuple)
        
        except Exception as e:
            print("{} generated an exception {}".format(foundTuple[0], e))

sortedFirmwareMacTuple = sorted(firmwareMacTuple, key = lambda x: x[2])

fieldWidth1 = 25
fieldWidth2 = 30
fieldWidth3 = 15

#Router results printout
formattedRouterFirmwareMacTuple = []
formattedFileServerFirmwareMacTuple = []

print("\n Routers: \n")

print(" {:<{}} | {:<{}} | {:<{}} | {:<{}}".format("LOCATION", "DEVICEN", "IP ADDRESS", "MAC ADDRESS", "FIRMWARE", fieldWidth3 = fieldWidth3, fieldWidth2 = fieldWidth2, fieldWidth1 = fieldWidth1))

for ferm in sortedFirmwareMacTuple:
    print(" {:<{}} | {:<{}} | {:<{}} | {:<{}}".format(ferm[2], ferm[0], ferm[1], ferm[4], ferm[3], fieldWidth1 = fieldWidth1, fieldWidth2 = fieldWidth2, fieldWidth3 = fieldWidth3))
    firmwareFormatString = ferm[3]
    firmwareFormatString = firmwareFormatString.split("\x00")[0]
    formatTuple = (ferm[2], ferm[0], ferm[1], ferm[4], firmwareFormatString, ferm[5], ferm[6])
    formattedRouterFirmwareMacTuple.append(formatTuple)

print("\nDo you want to also grab the current File Server info? Keyboard input y for Yes, n for No\n")

while True:
    if keyboard.is_pressed('y'):
        print("Please wait about 3 minutes for all file server values to download...\n")

        #File Server Firmware Check

        fileServerSecondaryFirmwareTuple = []
        fileServerSecondaryFirmwareMacTuple = []
        
        with ThreadPoolExecutor(max_workers = 16) as executor:
            futureTuple = {executor.submit(fileServerSecondaryFirmwareChecker, foundTuple): foundTuple for foundTuple in fileserverDevicesFoundTuple}
            
            for future in as_completed(futureTuple):
                foundTuple = futureTuple[future]
                
                try:
                    deviceFound, ipFound, locationFound, separatedFirmware, separatedMacAddress, newFileServerFirmware, newFileServerBootFirmware, newFileServerIndicator = future.result()
                    
                    newTuple = (deviceFound, ipFound, locationFound, separatedFirmware, separatedMacAddress, newFileServerFirmware, newFileServerBootFirmware, newFileServerIndicator) + estDateTimeTuple
                    fileServerSecondaryFirmwareTuple.append(newTuple)
                
                except Exception as e:
                    print("{} generated an exception {}".format(foundTuple[0], e))

        sortedFileServerFirmwareMacTuple = sorted(fileServerSecondaryFirmwareTuple, key = lambda x: x[2])
        
        #File Server results printout
        
        print("\n\nFileServers: \n")
        
        print(" {:<{fieldWidth3}} | {:<{fieldWidth2}} | {:<{fieldWidth1}} | {:<{fieldWidth1}} | {:<{fieldWidth2}} | {:<{fieldWidth2}} | {:<{fieldWidth2}} | {:<{fieldWidth2}}".format("LOCATION", "DEVICE NAME", "IP ADDRESS", "MAC ADDRESS", "NETWORK SERVER FIRMWARE", "FILE SERVER FIRMWARE", "SERVER BOOT FIRMWARE", "NEW SERVER INDICATOR", fieldWidth1 = fieldWidth1, fieldWidth2 = fieldWidth2, fieldWidth3 = fieldWidth3))
        
        for ferm in sortedFileServerFirmwareMacTuple:
            print(" {:<{}} | {:<{}} | {:<{}} | {:<{}} | {:<{}} | {:<{}} | {:<{}}".format(ferm[2], ferm[0], ferm[1], ferm[4], ferm[3], ferm[5], ferm[6], ferm[7], fieldWidth1 = fieldWidth1, fieldWidth2 = fieldWidth2, fieldWidth3 = fieldWidth3))
            formatTuple = (ferm[2], ferm[0], ferm[1], ferm[4], ferm[3], ferm[5], ferm[6], ferm[7], ferm[8], ferm[9])
            formattedFileServerFirmwareMacTuple.append(formatTuple)
        break

    if keyboard.is_pressed('n'):
        formatTuple = (" ", ) * 10
        formattedFileServerFirmwareMacTuple = [formatTuple]
        
        break
        
localFilePath = "RoutersFileServersFirmwareResults.xlsx"
sharedFilePath = "\\exampleTeamNetworkPath\\RoutersFileServersFirmwareResults\\RoutersFileServersFirmwareResults.xlsx"

routerTopRow = ["LOCATION", "DEVICE NAME", "IP ADDRESS", "MAC ADDRESS", "FIRMWARE", "DATE", "TIME"]
fileServerTopRow = ["LOCATION", "DEVICE NAME", "IP ADDRESS", "MAC ADDRESS", "NETWORK SERVER FIRMWARE", "FILE SERVER FIRMWARE", "SERVER BOOT FIRMWARE", "NEW SERVER INDICATOR", "DATE", "TIME"]

excelFirmwareAppend(localFilePath, sharedFilePath, routerTopRow, fileServerTopRow, formattedRouterFirmwareMacTuple, formattedFileServerFirmwareMacTuple)

print("\n\nSaved to local file path: {}".format(localFilePath))
print("Saved to shared file path: {}\n".format(sharedFilePath))


#Determining total time to run script below
timeEnd = time.time()

timeDiff = timeEnd - timeStart

print(f"The Router & FileServer Firmware Info program took: {timeDiff:.2f} seconds to run. ")

#End by requiring the user to input 'q' on their keyboard when they're finished viewing results.
print("\n\nPress 'q' key on keyboard to quit program. ")

while True:
    if keyboard.is_pressed('q'):
        print("The 'q' key was pressed. Exiting program... ")
        time.sleep(3)
        break    